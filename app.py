"""
DACTE PDF Extractor - Flask Web Application
Upload DACTE PDFs and download extracted data as formatted Excel (.xlsx)
"""

import os
import io
import tempfile
from datetime import datetime

from flask import Flask, render_template, request, send_file, jsonify
from werkzeug.utils import secure_filename
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from extractor import extract_from_pdf

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB limit
app.config['UPLOAD_FOLDER'] = tempfile.mkdtemp()

ALLOWED_EXTENSIONS = {'pdf'}


def allowed_file(filename):
    """Check if file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def generate_excel(data_list: list) -> io.BytesIO:
    """
    Generate a formatted Excel file from extracted DACTE data.
    
    Args:
        data_list: List of dicts with extracted fields
        
    Returns:
        BytesIO object containing the .xlsx file
    """
    # Create DataFrame with correct column order
    df = pd.DataFrame(data_list)
    df = df.rename(columns={
        'cte': 'CTE',
        'tipo': 'Tipo',
        'data_emissao': 'Data emissão',
        'planta': 'Planta',
        'valor': 'Valor',
        'valor_servico': 'Valor serviço',
        'conteiner': 'CONTEINER'
    })
    
    # Ensure correct column order
    columns = ['CTE', 'Tipo', 'Data emissão', 'Planta', 'Valor', 'Valor serviço', 'CONTEINER']
    for col in columns:
        if col not in df.columns:
            df[col] = ''
    df = df[columns]
    
    # Write to BytesIO
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='DACTEs')
    
    output.seek(0)
    
    # Now apply formatting with openpyxl
    wb = load_workbook(output)
    ws = wb.active
    
    # Header styling - dark blue background, white bold text
    header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='4472C4'),
        right=Side(style='thin', color='4472C4'),
        top=Side(style='thin', color='4472C4'),
        bottom=Side(style='thin', color='4472C4')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data cell styling
    data_font = Font(name='Calibri', size=11)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # Currency formatting for Valor (column E = 5) and Valor serviço (column F = 6)
    currency_format = 'R$ #,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=6):
        for cell in row:
            if cell.value is not None:
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Auto-fit column widths with currency-aware sizing
    # Column minimum widths to prevent ### display
    min_widths = {
        'A': 10,   # CTE
        'B': 10,   # Tipo
        'C': 16,   # Data emissão
        'D': 22,   # Planta
        'E': 22,   # Valor (R$ 1.603.426,71 = ~18 chars)
        'F': 20,   # Valor serviço (R$ 5.746,84 = ~14 chars)
        'G': 16,   # CONTEINER
    }
    
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value is not None:
                    # For currency columns, estimate formatted width
                    if column_letter in ('E', 'F') and isinstance(cell.value, (int, float)):
                        # Format: "R$ 1.234.567,89" — estimate display length
                        formatted = f"R$ {cell.value:,.2f}"
                        max_length = max(max_length, len(formatted))
                    else:
                        max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        # Use the larger of calculated width or minimum width
        min_w = min_widths.get(column_letter, 10)
        adjusted_width = max(max_length + 4, min_w)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    # Add autofilter
    ws.auto_filter.ref = ws.dimensions
    
    # Save to BytesIO
    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output


@app.route('/')
def index():
    """Serve the main upload page."""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle PDF file uploads, extract data, and return Excel download."""
    if 'files' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    files = request.files.getlist('files')
    
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    
    all_data = []
    errors = []
    
    for file in files:
        if not file or file.filename == '':
            continue
            
        if not allowed_file(file.filename):
            errors.append(f'{file.filename}: Formato não suportado (apenas PDF)')
            continue
        
        try:
            # Save temporarily
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Extract data
            extracted = extract_from_pdf(filepath)
            all_data.extend(extracted)
            
            # Clean up
            os.remove(filepath)
            
        except Exception as e:
            errors.append(f'{file.filename}: Erro na extração - {str(e)}')
    
    if not all_data:
        error_msg = 'Nenhum dado extraído dos PDFs.'
        if errors:
            error_msg += ' Erros: ' + '; '.join(errors)
        return jsonify({'error': error_msg}), 400
    
    # Generate Excel
    try:
        excel_file = generate_excel(all_data)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        download_name = f'DACTE_Extraidos_{timestamp}.xlsx'
        
        # Return both data summary and trigger download
        return send_file(
            excel_file,
            download_name=download_name,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Erro ao gerar Excel: {str(e)}'}), 500


@app.route('/extract', methods=['POST'])
def extract_preview():
    """Extract data and return JSON preview (without Excel download)."""
    if 'files' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    files = request.files.getlist('files')
    
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    
    all_data = []
    errors = []
    
    for file in files:
        if not file or file.filename == '':
            continue
            
        if not allowed_file(file.filename):
            errors.append(f'{file.filename}: Formato não suportado (apenas PDF)')
            continue
        
        try:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            extracted = extract_from_pdf(filepath)
            all_data.extend(extracted)
            
            os.remove(filepath)
            
        except Exception as e:
            errors.append(f'{file.filename}: Erro na extração - {str(e)}')
    
    return jsonify({
        'data': all_data,
        'count': len(all_data),
        'errors': errors
    })


if __name__ == '__main__':
    app.run(debug=True, port=5000)
